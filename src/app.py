import openpyxl
from tabulate import tabulate

xlsx_dataframe = openpyxl.load_workbook("Sheet 3.xlsx")

dataframe = xlsx_dataframe.active

data = []

for row in range(1, dataframe.max_row):

    _row = [row,]

    for col in dataframe.iter_cols(1, dataframe.max_column):

        _row.append(col[row].value)
    
    data.append(_row)

data = data[7:]
headers = [
    "Event Name",
    "Source/Medium of Session",
    "Sessions with Interaction",
    "Sessions",
    "Events per Session",
    "Sessions with Interaction per Active User",
    "Views per Session",
    "Bounce Rate",
    "Interaction Rate",
    "Average Interaction Time per Session"
]

print(tabulate(data, headers=headers))